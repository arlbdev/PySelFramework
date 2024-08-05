import openpyxl


book = openpyxl.load_workbook("C:\\Users\\Azi\\PycharmProject\\PythonTesting\\PySelFramework\\ExcelTestDataSample\\UsersSample.xlsx")
sheet = book.active

testData = {}

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestData2":
        for j in range(2, sheet.max_column + 1):
            testData[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(testData)
