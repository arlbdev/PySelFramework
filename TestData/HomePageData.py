import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstName": "Juan", "email": "juandelacruz@gmail.com", "gender": "Male"},
                          {"firstName": "Jane", "email": "janedoe@gmail.com", "gender": "Female"}
                          ]

    @staticmethod
    def getTestData(test_case_name):

        book = openpyxl.load_workbook(
            "C:\\Users\\Azi\\PycharmProject\\PythonTesting\\PySelFramework\\ExcelTestDataSample\\UsersSample.xlsx")
        sheet = book.active

        test_data = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    test_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [test_data]
